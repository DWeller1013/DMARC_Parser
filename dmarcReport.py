# -----------------------------------------------------------------------------
# DMARC Reporter
# -----------------------------------------------------------------------------
# Download all attachments in emails from DMARC subfolder. Unzip and extract
# all .zip and .gz files. Parse through .xml files and display data in an excel sheet.
# Organize and format excel sheet and email back to desired user.
# Eventually will set up as Crontab to send reports every Thursday morning.
# -----------------------------------------------------------------------------
# 250721(dmw) TODO: Look into using WHOIS through command line again... Not RDAP?
#			- Speed up processing through IPs. Grab OrgName, Country, etc... 
#
# 250620(dmw) Added SPF_Failures sheet. Investigate why SPF is failing.
# 250620(dmw) Added get_org_name function. RDAP lookup and added caching. Removed graphs/charts for now.
# 250612(dmw) Added PieChart with percentages pulling from TabularData.
# 250612(dmw) Added TabularData function.
# 250609(dmw) Added Email Reporting for parsed DMARC xml files.
# -----------------------------------------------------------------------------

import email
import datetime
from email.header import decode_header
from email.message import EmailMessage
import gzip
import imaplib
import mimetypes
import pickle
import os
import time
import shutil
import smtplib
import zipfile
import dns.resolver
import ipaddress
import socket
import requests
import re
import xml.etree.ElementTree as ET

import pandas as pd
import plotly.express as px
from dotenv import load_dotenv
from openpyxl import load_workbook
from ipwhois import IPWhois
from tqdm import tqdm
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image as XLImage

# -----------------------------------------------------------------------------
# Globals
CURRENT_DATE = datetime.datetime.now().strftime("%Y-%m-%d")
DNS_CACHE_FILE = "dns_cache.pkl"
HOST_CACHE_FILE = "host_cache.pkl"
GEO_CACHE_FILE = "geo_cache.pkl"
tqdm.pandas()

# -----------------------------------------------------------------------------
# Load environment variables from .env file
def load_config():
	load_dotenv() 
	# config returns a dictionary with the following keys:
	config = {
		"EMAIL": os.getenv("EMAIL"),
		"PASSWORD": os.getenv("PASSWORD"),
		"IMAP_SERVER": os.getenv("IMAP_SERVER"),
		"IMAP_SSL_PORT": int(os.getenv("IMAP_SSL_PORT")),
		"TO_EMAIL": os.getenv("TO_EMAIL"),
		"FROM_EMAIL": os.getenv("FROM_EMAIL"),
	}
	return config

# -----------------------------------------------------------------------------
# Establish a secure IMAP connection to the mail server and logs in
def connect_imap(server, port, email_addr, password):
	mail = imaplib.IMAP4_SSL(server, port) # Authenticated IMAP4_SSL object
	mail.login(email_addr, password)
	print("IMAP login successful.")
	return mail

# -----------------------------------------------------------------------------
# Searches for emails in the specified IMAP folder received within the last # of days
def search_recent_emails(mail, folder="INBOX", days=7):
	mail.select("DMARC") # Select the DMARC folder inside INBOX
	date_since = (datetime.datetime.now() -
				  datetime.timedelta(days=7)).strftime("%d-%b-%Y")
	status, message_numbers = mail.search(None, f'SINCE {date_since}')
	ids = message_numbers[0].split()
	return ids # Return a list of email IDs (as bytes) received within the date range

# -----------------------------------------------------------------------------
# Create a dictionary for saving attachments.
def make_save_dir(base="attachments"):
	global CURRENT_DATE
	save_dir = os.path.join("attachments", CURRENT_DATE)
	os.makedirs(save_dir, exist_ok=True)
	return save_dir # Full path to the directory for today's date

# -----------------------------------------------------------------------------
# Download and save attachments from a list of email IDs in the spcified directory
def save_attachments(mail, ids, save_dir):
	# For each email ID, fetch full email message with RFC822 protocol
	for num in ids:
		status, msg_data = mail.fetch(num, '(RFC822)')
		for response_part in msg_data:
			if isinstance(response_part, tuple):
				msg = email.message_from_bytes(response_part[1])
				for part in msg.walk(): # Traverse all the branches of the msg.
					if part.get_content_maintype() == 'multipart': # skip multipart containers (not attachments)
						continue
					if part.get('Content-Disposition') is None: # Check for 'Content-Disposition' header to identify attachments
						continue
					filename = part.get_filename()
					if filename: # if fileName is present, save attachment to save_dir.
						filepath = os.path.join(save_dir, filename)
						with open(filepath, 'wb') as f:
							f.write(part.get_payload(decode=True))
						print(f"Saved attachment: {filepath}")

# -----------------------------------------------------------------------------
# Unzip all .zip and .gz files found in the specified directory (save_dir)
# Place extracted files into subdirectory called 'unzipped'
def unzip_files(save_dir):
	unzipped_dir = os.path.join(save_dir, "unzipped")
	os.makedirs(unzipped_dir, exist_ok=True)
	for filename in os.listdir(save_dir): # Iterate over all files in save_dir
		file_path = os.path.join(save_dir, filename)
		if os.path.isdir(file_path): # If file is directory, skip
			continue
		if filename.lower().endswith(".zip"): # Extract contents to unzipped_dir
			try:
				with zipfile.ZipFile(file_path, 'r') as zip_ref:
					zip_ref.extractall(unzipped_dir)
				print(f"Unzipped {filename} to {unzipped_dir}")
			except Exception as e:
				print(f"Failed to unzip {filename}: {e}")
		elif filename.lower().endswith(".gz"): # Extract contents to unzipped_dir
			try:
				out_name = filename[:-3] # Remove .gz
				out_path = os.path.join(unzipped_dir, out_name)
				with gzip.open(file_path, 'rb') as f_in:
					with open(out_path, 'wb') as f_out:
						shutil.copyfileobj(f_in, f_out)
				print(f"Unzipped {filename} to {out_path}")
			except Exception as e:
				print(f"Failed to unzip {filename}: {e}")

# -----------------------------------------------------------------------------
# Format the excel sheets 
def formatSheets(excel_path):
	wb = load_workbook(excel_path)

	fixed_width_columns = {
		'spf_record': 30,
		'spf_includes': 30,
		'spf_notes': 30
	}

	# Loop through every worksheet in the workbook
	for sheet_name in wb.sheetnames:
		ws = wb[sheet_name]
		print(f"Formatting sheet: {sheet_name}")
		header_row = [cell.value for cell in ws[1]]
		col_map ={header: get_column_letter(idx + 1) for idx, header in enumerate(header_row)}
		# Loop through all columns in the worksheet 
		for i, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row, max_col=ws.max_column), 1):
			#col_letter = col[0].column_letter # Get the Excel-style column letter
			col_letter = get_column_letter(i)
			header = col[0].value # Get the header value from the first row
			if header in fixed_width_columns:
				ws.column_dimensions[col_letter].width = fixed_width_columns[header]
			else: 
				max_length = 0 # Track the maximum content length for that column
				# Loop through each cell in the column
				for cell in col:
					if header == 'source_ip' or header == 'envelope_to':
						cell.alignment = Alignment(horizontal='left')
					else:
						cell.alignment = Alignment(horizontal='center')

					# Calculate content length for auto-width
					cell_value = str(cell.value) if cell.value is not None else ''
					max_length = max(max_length, len(cell_value))

				# Set column width with some padding
				ws.column_dimensions[col_letter].width = max_length + 1 

	# Save changes to file
	wb.save(excel_path)
	print(f"Formatting complete for all sheets in '{excel_path}'")

# -----------------------------------------------------------------------------
# Check if old_name sheet exists in workbook and replace with new_name.
def renameSheet(excel_path, old_name, new_name):
	wb = load_workbook(excel_path)
	if old_name in wb.sheetnames:
		ws = wb[old_name] # Access the worksheet by old_name
		ws.title = new_name # Rename the worksheet
		wb.save(excel_path) # Save the workbook and apply changes
		print(f"Renamed sheet '{old_name}' to '{new_name}'")
	else:
		print(f"Sheet '{old_name}' not found, cannot rename.")


# -----------------------------------------------------------------------------
# Parse all DMARC XML files in the specified directory, extract relevant information
# and write the results to an excel file (overwrites if preexisting) 
def parse_dmarc_directory(unzipped_dir, report_dir, date_str):
	
	os.makedirs(report_dir, exist_ok=True)
	all_records = [] # Create dictionary 

	for filename in os.listdir(unzipped_dir): # Iterate through all files in unzipped_dir
		file_path = os.path.join(unzipped_dir, filename)
		if os.path.isdir(file_path): # Skip subdirectories 
			continue
		if filename.lower().endswith(".xml"):
			try:
				tree = ET.parse(file_path)
				root = tree.getroot()
				for record in root.findall(".//record"):
					row = {
						'envelope_to': record.findtext('./identifiers/envelope_to'), # Address sent to 
						'source_ip': record.findtext('./row/source_ip'), # IP address source of DMARC record
						'count': record.findtext('./row/count'), # Number of messages for this record
						'disposition': record.findtext('./row/policy_evaluated/disposition'), # DMARC Policy result (None - no action, quarantine - move to spam, reject - rejected the email)
						'dkim_result': record.findtext('./row/policy_evaluated/dkim'), # DKIM Evaluation result - Check if message is signed using a valid key and if the domain in the DKIM signature
																					   # (d=) or SPF record matches the domain in the "From" address of the email
						'spf_result': record.findtext('./row/policy_evaluated/spf'), # SPF Evaluation Result - Checks if the email server sending the message is authorized by the domain to send emails on its behalf
						'header_from': record.findtext('./identifiers/header_from'),
						'spf_checked_domain': '',
						'spf_checked_result': '',
						'spf_for_header_from': '',
					}
					spf_auths = record.findall('./auth_results/spf')
					spf_for_header_from = None
					for spf in spf_auths:
						this_domain = spf.findtext('domain')
						this_result = spf.findtext('result')
						if this_domain and this_result:
							if this_domain == row['header_from']:
								spf_for_header_from = this_result
							row['spf_checked_domain'] = this_domain
							row['spf_checked_result'] = this_result
					row['spf_for_header_from'] = spf_for_header_from
					all_records.append(row) # Append each record as dictionary to all_records
			except Exception as e:
				print(f"Failed to parse {filename}: {e}")

	if all_records:
		df = pd.DataFrame(all_records) # Converts all_records to a DataFrame 
		excel_path = os.path.join(report_dir, f"dmarc_report_{date_str}.xlsx")
		df.to_excel(excel_path, index=False) # Write DataFrame to Excel file.
		renameSheet(excel_path, 'Sheet1', 'All Data')
		print(f"\nDMARC report written to {excel_path}")
		return excel_path
	else:
		print("No DMARC records were found.")
		return None

# -----------------------------------------------------------------------------
# 250620(dmw) Check each source_ip for the DNS org's name using RDAP Lookup (WHOIS) to query the IP.
def get_org_name(ip, cache):
	if ip in cache:
		return cache[ip]

	try:
		response = requests.get(f"http://ipinfo.io/{ip}/json", timeout=3)
		data = response.json()
		print(f"ORG {ip}: {data}")
		org = data.get("org", "Unknown")
	except Exception:
		print(f"ORG ERROR {ip}: {e}")
		org = "Unknown"

	cache[ip] = org
	return org

# -----------------------------------------------------------------------------
# Lookup Host Name for each IP.
def get_hostname(ip, cache):
	if ip in cache:
		return cache[ip]

	try:
		hostname = socket.gethostbyaddr(ip)[0]
	except Exception:
		hostname = ""
	
	cache[ip] = hostname
	return hostname

# -----------------------------------------------------------------------------
# Lookup Geolocation for each IP.
def get_geolocation(ip, cache):
	if ip in cache:
		return cache[ip]
	
	try:
		response = requests.get(f"https://ipinfo.io/{ip}/json")
		data = response.json()
		city = data.get("city", "")
		region = data.get("region", "")
		country = data.get("country", "")
		geo_str = ", ".join([x for x in [city, region, country] if x]) 
	except Exception:
		geo_str = ""

	cache[ip] = geo_str
	return geo_str

# -----------------------------------------------------------------------------
def get_ipinfo(ip, org_cache, host_cache, geo_cache):
	try:
		#r = requests.get(f"https://ipinfo.io/{ip}/json", timeout=3) #, verify = False)
		url = f"http://ip-api.com/json/{ip}?fields=status,message,org,reverse,country"
		r = requests.get(url, timeout=3)
		data = r.json()
		#obj = IPWhois(ip)
		#results = obj.lookup_rdap(depth=1)

		if data.get("status") == "success":
			# Org
			org = data.get('org', 'Unknown')
			if org.startswith("AS") and " " in org:
				org = org.split(" ", 1)[1]
			org_cache[ip] = org
			#org = results.get('asn_description', 'Unknown')
			#org_cache[ip] = org if org else "Unknown"

			# Hostname
			host_cache[ip] = data.get('reverse', '')
			#host_cache[ip] = data.get('hostname', '')
			#hostname = results.get('network', {}).get('name', '')
			#host_cache[ip] = hostname if hostname else ""

			# Geolocation
			geo_cache[ip] = data.get('country', '')
			#country = results.get('network', {}).get('country', '')
			#geo_cache[ip] = country if country else ""
		
		else:
			print(f"Unsuccessful API Call...")
			org_cache[ip] = "Unknown"
			host_cache[ip] = ""
			geo_cache[ip] = ""

	except Exception:
		org_cache[ip] = "Unknown"
		host_cache[ip] = ""
		geo_cache[ip] = ""

# -----------------------------------------------------------------------------
# Load the caches
def load_cache(filename):
	try:
		with open(filename, "rb") as f:
			return pickle.load(f)
	except Exception:
		return {}	

# -----------------------------------------------------------------------------
# Save the caches
def save_cache(cache, filename):
	try:
		with open(filename, "wb") as f:
			pickle.dump(cache, f)
	except Exception:
		return {}

# -----------------------------------------------------------------------------
# Read all data for each row and organize into a more readable format.
def organizeData(excel_path):
	
	# Load cache files
	dns_cache = load_cache(DNS_CACHE_FILE)
	host_cache = load_cache(HOST_CACHE_FILE)
	geo_cache = load_cache(GEO_CACHE_FILE)

	try:
		# Read the data from the specified sheet
		df = pd.read_excel(excel_path)

		print(f"Starting IP Lookups...")

		# Deduplicate IPs
		unique_ips = df['source_ip'].dropna().unique()
		tqdm.pandas()

		for ip in tqdm(unique_ips, desc="Preprocessing unique IPs"):
			if ip in dns_cache and ip in host_cache and ip in geo_cache:
				continue # Already cached, skip.
			get_ipinfo(ip, dns_cache, host_cache, geo_cache)
			time.sleep(1.5) # ~40 requests per minute, safe for ip-api.com
			#get_org_name(ip, dns_cache)
			#get_hostname(ip, host_cache)
			#get_geolocation(ip, geo_cache)

		# Assign columns from caches
		if 'source_dns' not in df.columns:
			df.insert(df.columns.get_loc('source_ip') + 1,
						'source_dns',
						df['source_ip'].map(dns_cache))
		if 'source_host' not in df.columns:
			df.insert(df.columns.get_loc('source_dns') + 1,
						'source_host',
						df['source_ip'].map(host_cache))
		if 'source_geo' not in df.columns:
			df.insert(df.columns.get_loc('source_host') + 1,
						'source_geo',
						df['source_ip'].map(geo_cache))

		#print(f"Source DNS Lookups...")
		## Add source_dns if missing
		#if 'source_dns' not in df.columns:
		#	df.insert(df.columns.get_loc('source_ip') + 1,
		#		'source_dns',
		#		df['source_ip'].progress_apply(lambda ip: get_org_name(ip, dns_cache))
		#	)
		#
		#print(f"Host Name Lookups...")
		## Add host name if missing
		#if 'source_host' not in df.columns:
		#	df.insert(df.columns.get_loc('source_dns') + 1,
		#		'source_host',
		#		df['source_ip'].progress_apply(lambda ip: get_hostname(ip, host_cache))
		#	)
	
		#print(f"Geolocation Lookups...")
		## Add geolocation if missing
		#if 'source_geo' not in df.columns:
		#	df.insert(df.columns.get_loc('source_host') + 1,
		#		'source_geo',
		#		df['source_ip'].progress_apply(lambda ip: get_geolocation(ip, geo_cache))
		#	)

		print(f"Checking auth_status...")
		# Add auth_status
		df['auth_status'] = df.apply(
			lambda row: 'Authenticated' if row['dkim_result'] == 'pass' or row['spf_result'] == 'pass' else 'Failed',
			axis=1
		)

		df['spf_result'] = df['spf_result'].fillna('')
		df['spf_fail_sort'] = (df['spf_result'].str.lower() == 'fail').astype(int)
		df = df.sort_values(by='spf_fail_sort', ascending=False).drop(columns='spf_fail_sort')

		# Write the organized summary to the new sheet
		with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
			df.to_excel(writer, sheet_name="Organized_Data", index=False)

		print(f"Table created on sheet 'Organized Data' in '{excel_path}'.")
		
	except FileNotFoundError:
		print(f"Error: the file '{excel_file}' was not found.")
	except Exception as e:
		print(f"An error occurred: {e}")

	# Save cache to file (persist new lookups)
	save_cache(dns_cache, DNS_CACHE_FILE)
	save_cache(host_cache, HOST_CACHE_FILE)
	save_cache(geo_cache, GEO_CACHE_FILE)

	# Save cache to file (persist new lookups)
	#with open(CACHE_FILE, "wb") as f:
	#	pickle.dump(dns_cache, f)
	#
	## Save host cache to file (persist new lookups)
	#with open(HOST_CACHE_FILE, "wb") as f:
	#	pickle.dump(host_cache, f)
	#
	## Save geo cache to file (persist new lookups)
	#with open(GEO_CACHE_FILE, "wb") as f:
	#	pickle.dump(geo_cache, f)

# -----------------------------------------------------------------------------
# Send the DMARC Excel report as an email attachment
def emailReport():
	global CURRENT_DATE
	# Load the config, .env file contains all data to send email
	config = load_config()
	smtp_server = os.getenv("SMTP_SERVER")
	smtp_port = int(os.getenv("SMTP_PORT", 587)) # Default to 587 if SMTP_PORT is not present.
	from_email = config["FROM_EMAIL"]
	to_email = config["TO_EMAIL"]
	email = config["EMAIL"]
	password = config["PASSWORD"]

	# Prepare date strings for the report period 
	prev_date = (datetime.datetime.now() - datetime.timedelta(days=7)).strftime("%Y-%m-%d")
	report_dir = os.path.join(os.getcwd(), "Dmarc_Reports")
	excel_filename = f"dmarc_report_{CURRENT_DATE}.xlsx"
	excel_path = os.path.join(report_dir, excel_filename)

	# Check if the report file exists
	if not os.path.exists(excel_path):
		print(f"Report file not found: {excel_path}")
		return

	# Construct email message
	msg = EmailMessage()
	msg["Subject"] = f"DMARC Report - {prev_date} - {CURRENT_DATE}."
	msg["To"] = to_email
	msg["From"] = from_email
	msg.set_content(f"Attached is the DMARC report for:\n{prev_date} - {CURRENT_DATE}")

	# Determine the MIME type for the attachment
	ctype, encoding = mimetypes.guess_type(excel_path)
	if ctype is None or encoding is not None:
		ctype = "application/octet-stream" # Default for binary files with unknown formats
	maintype, subtype = ctype.split("/", 1)
	with open(excel_path, "rb") as f: # rb = Read/Binary
		msg.add_attachment(f.read(), maintype=maintype, subtype=subtype, filename=excel_filename)

	try:
		with smtplib.SMTP(smtp_server, smtp_port) as server:
			server.starttls()
			#server.login(email, password)
			server.send_message(msg)
			server.quit()
		print(f"Report emailed to {to_email} from {from_email}")
	except Exception as e:
		print(f"Failed to send email: {e}")

# -----------------------------------------------------------------------------
def generatePlotlyChart(df, html_path="dmarc_plotly_report.html", image_path="dmarc_plot.png"):
	
	df.columns = df.columns.str.strip()

	# Metrics
	metrics = {
		'DMARC': 'DMARC Compliance',
		'SPF': 'SPF',
		'DKIM': 'DKIM'
	}

	chart_files = []

	for label, col in metrics.items():
		# Group by pass/fail, sum email volume
		summary = df.groupby(col)['Email volume'].sum().reset_index()
		# Calculate passrate
		total = summary['Email volume'].sum()
		pass_row = summary[summary[col]].str.lower() == 'pass'


	#fig.show()
	fig.write_html(html_path)
	fig.write_image(image_path)
	print(f"Plotly chart saved as {html_path} and image as {image_path}")
	return html_path, image_path

# -----------------------------------------------------------------------------
def chartData(excel_path):

	wb = load_workbook(excel_path)
	ws = wb['Tabular Data']

	if "Charts" not in wb.sheetnames:
		chart_ws = wb.create_sheet("Charts")
	else:
		chart_ws = wb["Charts"]

	# ---------------------------------
	# Pie Chart for overall DMARC Pass/Fail
	total_pass = sum((v[0] for v in ws.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True) if isinstance(v[0], (int, float))), 0)
	total_fail = sum((v[0] for v in ws.iter_rows(min_row=2, min_col=4, max_col=4, values_only=True) if isinstance(v[0], (int, float))), 0)
	
	# Summary data for pie chart
	chart_ws["A25"] = "Pass"
	chart_ws["A26"] = "Fail"
	chart_ws["B25"] = total_pass
	chart_ws["B26"] = total_fail

	total = total_pass + total_fail
	pass_pct = total_pass / total if total else 0
	fail_pct = total_fail / total if total else 0

	chart_ws["C24"] = "Percent"
	chart_ws["C25"] = pass_pct
	chart_ws["C26"] = fail_pct

	chart_ws["C25"].number_format = '0.00%'
	chart_ws["C26"].number_format = '0.00%'

	pie = PieChart()
	pie.title = "Overall DMARC Pass/Fail"
	labels = Reference(chart_ws, min_col=1, min_row=25, max_row=26)
	data = Reference(chart_ws, min_col=2, min_row=25, max_row=26)
	pie.add_data(data, titles_from_data=False)
	pie.set_categories(labels)
	chart_ws.add_chart(pie, "A2")
	# ---------------------------------

	wb.save(excel_path)

# -----------------------------------------------------------------------------
# Present data in tabular format similar to Google's DMARC Example 
def tabularData(excel_path):
	
	df = pd.read_excel(excel_path, sheet_name = "Organized_Data")
		
	# Group by source_ip
	grouped = df.groupby(['header_from', 'source_ip'])
	#grouped = df.groupby('source_ip')

	summary_data = []
	for (header_from, ip), group in grouped:
	#for ip, group in grouped:
		volume = group['count'].sum()

		# DMARC pass/fail (previously calculated 'auth_status')
		dmarc_pass = ((group['auth_status'] == 'Authenticated') * group['count']).sum()
		dmarc_fail = ((group['auth_status'] == 'Failed') * group['count']).sum()
		dmarc_rate = f"{(dmarc_pass / volume * 100):.2f}%" if volume else "0.00%"

		# SPF Pass is only for header_from alignment
		spf_for_header_from_pass = ((group['spf_for_header_from'] == 'pass' ) * group['count']).sum()
		spf_for_header_from_fail = ((group['spf_for_header_from'] != 'pass' ) * group['count']).sum()
		spf_for_header_from_rate = f"{(spf_for_header_from_pass / volume * 100):.2f}%" if volume else "0.00%"

		# SPF Counts
		#spf_pass = ((group['spf_result'] == 'pass') * group['count']).sum()
		#spf_fail = ((group['spf_result'] == 'fail') * group['count']).sum()
		#spf_rate = f"{(spf_pass / volume * 100):.2f}%" if volume else "0.00%"

		# DKIM Counts
		dkim_pass = ((group['dkim_result'] == 'pass') * group['count']).sum()
		dkim_fail = ((group['dkim_result'] == 'fail') * group['count']).sum()
		dkim_rate = f"{(dkim_pass / volume * 100):.2f}%" if volume else "0.00%"

		summary_data.append([
			header_from, ip, volume,
			dmarc_pass, dmarc_fail, dmarc_rate,
			spf_for_header_from_pass, spf_for_header_from_fail, spf_for_header_from_rate, 
			dkim_pass, dkim_fail, dkim_rate
		])

		#summary_data.append([
		#	ip, volume,
		#	dmarc_pass, dmarc_fail, dmarc_rate,
		#	spf_pass, spf_fail, spf_rate, 
		#	dkim_pass, dkim_fail, dkim_rate
		#])

	columns = [
		'Header From', 'SourceIP Address', 'Email volume',
		'DMARC Pass', 'DMARC Fail', 'DMARC Rate',
		'SPF (aligned) Pass', 'SPF (aligned) Fail', 'SPF (aligned) Rate', 
		'DKIM Pass', 'DKIM Fail', 'DKIM Rate'
	]

#	columns = [
#		'IP Address', 'Email volume',
#		'DMARC Pass', 'DMARC Fail', 'DMARC Rate',
#		'SPF Pass', 'SPF Fail', 'SPF Rate', 
#		'DKIM Pass', 'DKIM Fail', 'DKIM Rate'
#	]

	summary_df = pd.DataFrame(summary_data, columns=columns)
	summary_df = summary_df.sort_values(by='SPF (aligned) Fail', ascending=False) # Sort by highest SPF fail rate

	# Save to Excel (append as new sheet)
	with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
		summary_df.to_excel(writer, sheet_name="Tabular Data", index=False)

	# Style the tabular data
	wb = load_workbook(excel_path)
	ws = wb["Tabular Data"]

	# Header formatting
	yellow = PatternFill("solid", fgColor="FFF475")
	bold = Font(bold=True)
	center = Alignment(horizontal="center", vertical="center")
	border = Border(bottom=Side(style="thin"), top=Side(style="thin"),
					  left=Side(style="thin"), right=Side(style="thin"))
	
	# First header row (manually set)
	#ws.merge_cells('A1:A2')
	#ws.merge_cells('B1:B2')
	#ws['A1'] = "IP Address"
	#ws['B1'] = "Email volume"

	#ws.merge_cells('C1:E1')
	#ws['C1'] = "DMARC Compliance"

	#ws.merge_cells('F1:H1')
	#ws['F1'] = "SPF"
	#
	#ws.merge_cells('I1:K1')
	#ws['I1'] = "DKIM"

	## Second header row (sub columns)
	#ws['C2'] = "Pass"
	#ws['D2'] = "Fail"
	#ws['E2'] = "Rate"
	#
	#ws['F2'] = "Pass"
	#ws['G2'] = "Fail"
	#ws['H2'] = "Rate"
	#
	#ws['I2'] = "Pass"
	#ws['J2'] = "Fail"
	#ws['K2'] = "Rate"

	# Style headers
	#for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=ws.max_column):
	#	for cell in row:
	#		cell.fill = yellow
	#		cell.font = bold
	#		cell.alignment = Alignment(horizontal='center')
	#		cell.border = border

	# Header styling
	for cell in ws[1]:
		cell.font = bold
		cell.fill = yellow
		cell.alignment = center
		cell.border = border

	# Set column widths automatically
	for i, col in enumerate(ws.columns, 1):
		max_length = 0
		col_letter = get_column_letter(i)
		#col_letter = col[0].column_letter
		for cell in col:
			try:
				if cell.value:
					max_length = max(max_length, len(str(cell.value)))
			except:
				pass
		ws.column_dimensions[col_letter].width = max_length + 4 

	# Freeze header and save
	ws.freeze_panes = "A2"
	wb.save(excel_path)
	print(f"Tabular DMARC summary added to '{excel_path}'.")

# -----------------------------------------------------------------------------
# Get the SPF Record for a given domain (first v=spf1 record found)
def get_spf_record(domain):
	try: 
		answers = dns.resolver.resolve(domain, 'TXT')
		for rdata in answers:
			# Concatenate all strings (TXT records sometimes split)
			txt = "".join([s.decode("utf-8") if isinstance(s, bytes) else s for s in rdata.strings])
			if txt.startswith('v=spf1'):
				return txt
	except Exception:
		pass
	return None

# -----------------------------------------------------------------------------
# Extract all include: return a list of domains
def extract_includes(spf_record):
	if not spf_record:
		return []
	# Find all "include:domain" entries in spf_record
	return re.findall(r'include:([^\s]+)', spf_record)

# -----------------------------------------------------------------------------
# Recursively check if IP is allowed by the SPF record of the domain.
# Limits recursion and DNS lookups to 10 (per SPF RFC)
def ip_in_spf(ip, domain, lookup_count=0, max_lookups=10, checked_domains=None):
	if checked_domains is None:
		checked_domains = set()
	# Stop if DNS lookup limit eached or if domain was already checked.
	if lookup_count >= max_lookups or domain in checked_domains:
		return False

	checked_domains.add(domain)
	spf_record = get_spf_record(domain)
	if not spf_record:
		return False

	try:
		ip_obj = ipaddress.ip_address(ip)
	except Exception:
		return False

	for part in spf_record.split():
		# Check direct IPv4 entry
		if part.startswith("ip4:") and isinstance(ip_obj, ipaddress.IPv4Address):
			try:
				if '/' in part[4:]:
					net = ipaddress.IPv4Network(part[4:], strict=False)
					if ip_obj in net:
						return True
					else:
						if ip_obj == ipaddress.IPv4Address(part[4:]):
							return True
			except Exception:
				continue
		# Check direct IPv6 entry
		elif part.startswith("ip6:") and isinstance(ip_obj, ipaddress.IPv6Address):
			try:
				if '/' in part[4:]:
					net = ipaddress.IPv6Network(part[4:], strict=False)
					if ip_obj in net:
						return True
					else:
						if ip_obj == ipaddress.IPv6Address(part[4:]):
							return True
			except Exception:
				continue
		# Recurse into include: domains
		elif part.startswith("include:"):
			included_domain = part[len("include:") :]
			# Recursively check included domain, imcrementing lookup count
			if ip_in_spf(ip, included_domain, lookup_count+1, max_lookups, checked_domains):
				return True

	return False

# -----------------------------------------------------------------------------
# Investigate an SPF failure row: get SPF_record, includes, check IP Auth and summarize findings
def investigate_spf_failure(row):
	investigation = {
		"ip_in_spf": None,
		"spf_record": None,
		"spf_includes": None,
		"spf_notes": "",
		"why_spf_failed": ""
	}

	# Use header_from for SPF Checks
	domain = row.get("header_from", "")
	if not domain or pd.isna(domain):
		investigation["spf_notes"] = "No header_from domain found."
		investigation["why_spf_failed"] = "The email did not include a valid sender domain (header_from), so SPF could not be checked."
		return investigation

	spf_record = get_spf_record(domain)
	investigation["spf_record"] = spf_record if spf_record else "No SPF record found"
	includes = extract_includes(spf_record) if spf_record else []
	investigation["spf_includes"] = ", ".join(includes) if includes else "None"

	ip = row.get("source_ip", "")
	in_spf = ip_in_spf(ip, domain) if spf_record else False
	investigation["ip_in_spf"] = in_spf

	# Summary ntes for the results.
	if spf_record is None:
		investigation["spf_notes"] = f"No SPF record present for domain {domain}."
		investigation["why_spf_failed"] = (
			f"No SPF record found for {domain}. Receiving servers can't verify if this IP is allowed. "
			"Add an SPF record to your DNS to specify authorized senders."
		)
	elif in_spf:
		investigation["spf_notes"] = f"Source IP is directly allowed in SPF record for {domain}."
		investigation["why_spf_failed"] = (
			f"The sending IP ({ip}) is allowed in your SPF record, but another issue may be causing the failure."
		)
	elif includes:
		investigation["spf_notes"] = (
			f"Source IP may be authorized via include(s) for {domain} (Manual review required): {investigation['spf_includes']}."
		)
		investigation["why_spf_failed"] = (
			f"The sending IP ({ip}) is not directly listed in your SPF record for {domain}. "
			"Check your SPF 'include:' mechanisms and make sure this IP is authorized."
		)
	elif row.get("spf_checked_domain","") and row.get("spf_checked_domain", "") != domain and row.get("spf_checked_result", "") == "pass":
		investigation["spf_notes"] += f" SPF Passed for non-aligned domain ({row.get('spf_checked_domain','')}), likely due to forwarding."
		investigation["why_spf_failed"] = (
			f"Your message was likely forwarded. SPF passed for the forwarder's domain ({row.get('spf_checked_domain','')}), "
			f"but not for your domain ({domain}). This can happen with email forwarding."
		)
	else:
		investigation["spf_notes"] = (
			f"Source IP is not authorized in SPF for {domain}; SPF likely fails due to missing include or direct ip4/ip6 entry."
		)
		investigation["why_spf_failed"] = (
			f"The sending IP ({ip}) is not listed in your SPF record. "
			"If this is a legitimate sender, update your SPF record."
		)

	# Forwarding / alignment logic
	spf_checked_domain = row.get("spf_checked_domain","")
	spf_checked_result = row.get("spf_checked_result","")
	
	if (
		row.get("spf_result", "").lower() == "fail"
		and spf_checked_domain
		and spf_checked_domain != domain
		and spf_checked_result == "pass"
	):
		investigation["spf_notes"] += f" SPF passed for non-aligned domains ({spf_checked_domain})"
		investigation["why_spf_failed"] += (
			f" Your message was likely forwarded. SPF passed for the forwarder's domain ({spf_checked_domain}), "
			f"but not for your domain ({domain}). This can happen with email forwarding."
		)

	return investigation

# -----------------------------------------------------------------------------
# Main SPF Failure investigation routine. Create new sheet with all SPF failures and investigation results.
def spfFailures(excel_path):

	# Read Organized_Data for parsed DMARC records and filter for SPF failures only
	df = pd.read_excel(excel_path, sheet_name="Organized_Data")
	spf_failures = df[df['spf_result'].str.lower() == 'fail'].copy()
	if spf_failures.empty:
		print("No SPF failures found.")
		return
	
	print(f"Investigating {len(spf_failures)} SPF failures...")

	# Apply investigation to each row, adding new columns for SPF info
	investigation_results = spf_failures.progress_apply(investigate_spf_failure, axis=1, result_type='expand')
	# Combine original SPF failure data with investigation columns
	spf_failures_subset = spf_failures[["envelope_to", "source_ip", "source_dns", "spf_result", "count"]].reset_index(drop=True)
	result_df = pd.concat([spf_failures_subset, investigation_results.reset_index(drop=True)], axis=1)

	# Sort by ip_in_spf True
	result_df['ip_in_spf'] = result_df['ip_in_spf'].astype(bool)
	result_df = result_df.sort_values('ip_in_spf', ascending=False).reset_index(drop=True)

	# Save the investigation results to a new worksheet 'SPF_Failures'.
	with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
		result_df.to_excel(writer, sheet_name="SPF_Failures", index=False)

	# Adjust column widths for spf_record, spf_includes, spf_notes
	wb = load_workbook(excel_path)
	ws = wb["SPF_Failures"]
	col_names = {cell.value: idx+1 for idx, cell in enumerate(ws[1])}
	for col in ["spf_record", "spf_includes", "spf_notes"]:
		if col in col_names:
			ws.column_dimensions[get_column_letter(col_names[col])].width = 40
	wb.save(excel_path)

	print(f"SPF investigation sheet created as 'SPF_Failures' in {excel_path}.")

# -----------------------------------------------------------------------------
# -----------------------------------------------------------------------------
# Main execution function for the DMARC processing workflow
def main():

	global CURRENT_DATE

	# Load config to open email and download requested files
	config = load_config()
	print("EMAIL: " + config["EMAIL"])
	print("TO EMAIL: " + config["TO_EMAIL"])
	print("IMAP_SERVER: " + config["IMAP_SERVER"])
	print("IMAP_SSL_PORT: " + str(config["IMAP_SSL_PORT"]))

	try:
		# Connect to the IMAP server and authenticate.
		mail = connect_imap(
				config["IMAP_SERVER"],
				config["IMAP_SSL_PORT"],
				config["EMAIL"],
				config["PASSWORD"]
		)
		
		# Check last 7 days of emails in DMARC folder
		ids = search_recent_emails(mail, folder="DMARC", days=7)
		save_dir = make_save_dir(base="attachments")
		
		# Save and unzip all attachments
		save_attachments(mail, ids, save_dir)
		unzip_files(save_dir)
		
		# Grab current date and parse directory for report generation
		unzipped_dir = os.path.join(save_dir, "unzipped")
		report_dir = os.path.join(os.getcwd(), "Dmarc_Reports")
		excel_path = parse_dmarc_directory(unzipped_dir, report_dir, CURRENT_DATE)
		
		organizeData(excel_path)
		tabularData(excel_path)
		spfFailures(excel_path)
		formatSheets(excel_path)
		
		df_tabular = pd.read_excel(excel_path, sheet_name="Tabular Data")

		# Send email based on .env values
		emailReport()

		# Logout from IMAP server
		mail.logout()
		print("IMAP logout successful.")

	except Exception as e:
		print(f"IMAP login failed: {e}")

# -----------------------------------------------------------------------------

if __name__ == "__main__":
	main()

# -----------------------------------------------------------------------------
